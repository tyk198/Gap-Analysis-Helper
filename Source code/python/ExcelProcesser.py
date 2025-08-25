import pandas as pd
import numpy as np
from typing import Union, List, Optional
from openpyxl import load_workbook
import xlwings as xw
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from urllib.parse import quote
import os 
from openpyxl.worksheet.hyperlink import Hyperlink

class ExcelProcesser:
    def __init__(self):
        """
        Initialize the ImageProcesser with data and settings.
        
        Args:
            data: The data to process
            settings: ImageProcesserSettings containing processing configuration
            raw_image_folder_path: Path to raw images provided by owner class
        """
    @staticmethod
    def read_and_copy_worksheet(wb):
        """
        Copies the 'TESTING' worksheet to a new 'copy data' worksheet
        in the specified Excel file, using openpyxl.

        Args:
            input_file (str): The path to the Excel file.

        Returns:
            tuple: target_worksheet
        """
        source_ws = wb['TESTING']
        #if 'copy data' in wb.sheetnames:
        #    wb.remove(wb['copy data'])
        target_ws = wb.copy_worksheet(source_ws)
        target_ws.title = 'copy data'
        return target_ws

    @staticmethod
    def create_table(target_ws):
        """
        Creates a table from the data range in the specified worksheet.

        Args:
            target_ws (openpyxl.worksheet.worksheet.Worksheet): The target worksheet.

        Returns:
            openpyxl.worksheet.worksheet.Worksheet: The worksheet with the table.
        """
        # Find the data range
        max_row = target_ws.max_row
        max_col = target_ws.max_column
        data_range = f"A1:{get_column_letter(max_col)}{max_row}"
        
        # Create table
        table = Table(displayName="MyNewTable", ref=data_range)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        target_ws.add_table(table)
        
        return target_ws

    @staticmethod
    def add_column_with_formula(worksheet, column_name, formula, formula_params=None):
        """
        Adds a new column with a formula to the worksheet.

        Args:
            worksheet (openpyxl.worksheet.worksheet.Worksheet): The target worksheet.
            column_name (str): Name of the new column.
            formula (str): The formula template.
            formula_params (dict, optional): Parameters to format the formula.
        """
        # Find the first empty column
        max_col = worksheet.max_column
        new_col = max_col + 1
        new_col_letter = get_column_letter(new_col)
        
        # Add column header
        worksheet[f"{new_col_letter}1"] = column_name
        
        # Apply formula to cells (starting from row 2)
        for row in range(2, worksheet.max_row + 1):
            cell_ref = f"{new_col_letter}{row}"
            if formula_params:
                # Replace placeholders in formula
                formatted_formula = formula
                for key, value in formula_params.items():
                    if key.startswith('col_'):
                        # Handle column references (e.g., col_H, col_D)
                        formatted_formula = formatted_formula.replace(
                            "{" + key + "}", 
                            f"{value}{row}"
                        )
                    else:
                        # Handle cell references (e.g., cell for MID function)
                        formatted_formula = formatted_formula.replace(
                            "{" + key + "}", 
                            f"{get_column_letter(value)}{row}"
                        )
            else:
                formatted_formula = formula
            
            worksheet[cell_ref] = formatted_formula

    @staticmethod
    def add_hyperlink_to_column(worksheet, image_absolute_path, excel_absolute_path, row_id):
        """
        Calculates a relative path and adds it as a hyperlink to a cell.

        Args:
            worksheet (openpyxl.worksheet.worksheet.Worksheet): The target worksheet.
            image_absolute_path (str): The full, absolute path to the target image file.
            excel_absolute_path (str): The full, absolute path to the Excel file being modified.
            row_id: The ID to match in the "ROW ID" column.
        """

        # --- Find Column Headers ---
        headers = {cell.value: cell.column_letter for cell in worksheet[1]}
        hyperlink_col = headers.get("Hyperlink")
        row_id_col = headers.get("ROW ID")

        if not hyperlink_col or not row_id_col:
            print("Error: 'Hyperlink' or 'ROW ID' column not found.")
            return

        # --- 1. Calculate the relative path ---
        # The starting point for the relative path is the directory containing the Excel file.
        excel_directory = os.path.dirname(excel_absolute_path)
        
        # Calculate the path from the Excel directory to the image.
        # This correctly handles ".." to go up directories if needed.
        relative_path = os.path.relpath(image_absolute_path, start=excel_directory)

        # --- 2. Format the relative path for an Excel hyperlink ---
        # Replace backslashes with forward slashes (universal for hyperlinks)
        path_with_forward_slashes = relative_path.replace(os.sep, '/')
        
        # URL-encode the path to handle spaces and other special characters safely
        encoded_path = quote(path_with_forward_slashes)

        # --- 3. Find the target row and add the hyperlink ---
        target_row = None
        for row in range(2, worksheet.max_row + 1):
            cell_value = worksheet[f"{row_id_col}{row}"].value
            if cell_value is not None and str(cell_value) == str(row_id):
                target_row = row
                break

        if target_row:
            cell = worksheet[f"{hyperlink_col}{target_row}"]
            cell.value = "View"
            cell.hyperlink = encoded_path
            cell.style = "Hyperlink"
            print(f"For row_id {row_id}:")
            print(f"  - Calculated Relative Path: {relative_path}")
            print(f"  - Added Hyperlink: {encoded_path}")
        else:
            print(f"Could not find row with ROW ID: {row_id}")






    @staticmethod
    def add_column_header(worksheet,header_name:str):
        """
        Adds an 'Image Hyperlink' column header to the right of the existing table in the worksheet.

        Args:
            worksheet (openpyxl.worksheet.worksheet.Worksheet): The target worksheet.
        """
        from openpyxl.utils import get_column_letter
        
        # Find the rightmost column of the table
        max_col = worksheet.max_column
        
        # Add 'Image Hyperlink' header to the next column
        new_col = max_col + 1
        new_col_letter = get_column_letter(new_col)
        worksheet[f"{new_col_letter}1"] = header_name
    
    
    @staticmethod
    def add_column_header(worksheet, header_name: str):
        """
        Adds a column header to the worksheet if it does not already exist.

        Args:
            worksheet (openpyxl.worksheet.worksheet.Worksheet): The target worksheet.
            header_name (str): The name of the header to check for and add.
        
        Returns:
            str: The column letter of the header (either the existing one or the newly created one).
        """
        
        # --- 1. Check if the header already exists in the first row ---
        for cell in worksheet[1]:
            if cell.value == header_name:
                print(f"Header '{header_name}' already exists in column {cell.column_letter}. No action taken.")
                return cell.column_letter  # Return the existing column letter

        # --- 2. If the loop completes, the header was not found, so add it ---
        # Find the rightmost column of the table
        max_col = worksheet.max_column
        
        # Add the header to the next available column
        new_col = max_col + 1
        new_col_letter = get_column_letter(new_col)
        worksheet[f"{new_col_letter}1"] = header_name
        print(f"Header '{header_name}' added to new column {new_col_letter}.")
        
        return new_col_letter # Return the new column letter